#!/usr/bin/env python 
# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from pathlib import Path
import win32com.client as win32
from tkinter import *
from tkinter.ttk import *
import tkinter.messagebox
import windnd


def xls_save_as(file_path):
    new_file_path = file_path.replace(Path(file_path).suffix, r'.xlsx')
    print(new_file_path)
    excel = win32.gencache.EnsureDispatch('excel.application')
    pro = excel.Workbooks.Open(file_path)  # 打开要转换的excel
    pro.SaveAs(new_file_path, FileFormat=51)  # 另存为xlsx格式
    pro.Close()
    excel.Application.Quit()
    return new_file_path


def get_material_no(string):
    if string == None:
        return None
    else:
        L = re.findall(r'[HR][0-9]{8}', string)
        if L:
            return L[0]
        else:
            return None


def get_file_path(files):
    global e1
    msg = '\n'.join((item.decode('gbk') for item in files))
    e1.delete(0, END)
    e1.insert("insert", msg)


def entry1_clear(event):
    global e1
    e1.delete(0, END)


def entry2_clear(event):
    global e2
    e2.delete(0, END)


def split_string_part(string):
    """
    将数字、字符以及r'/'拆分
    例如: 输入r'100nF/50V'，返回['100', 'nF', '/', '50', 'V']
    :param string:
    :return: 返回列表
    """
    return re.findall(r'[a-zA-Z]+', string)


def get_device_type(string):
    """
    获取元器件类型
    例如: 输入r'DZ1'，返回r'DZ'
    :param string:
    :return:返回元器件类型
    """
    return split_string_part(string)[0]


def get_device_count(string):
    device_type = get_device_type(string)
    if device_type == 'R':
        if get_material_no(string) != None:
            return string.count(device_type) - 1
        else:
            return string.count(device_type)
    else:
        return string.count(device_type)


def get_recoil_of_material_row(material_no, sheet):
    for hang in range(2, sheet.max_row + 1):
        if sheet.cell(hang, 1).value == material_no:
            if sheet.cell(hang, 4).value != None:
                return hang
    return None


root = Tk()
root.title('开料单助手_DEMO')
# root.geometry('400x100')
l1 = Label(root, text='料单文件路径')
l1.grid(row=0, column=0)

tStringVar1 = StringVar()
tStringVar1.set('请输入料单文件的路径或拖动料单文件进此框')
e1 = Entry(root, textvariable=tStringVar1, width=40)
e1.grid(row=0, column=1)
e1.bind('<Triple-Button-1>', entry1_clear)

l2 = Label(root, text='领取数量')
l2.grid(row=1, column=0)

tStringVar2 = StringVar()
tStringVar2.set('15')
e2 = Entry(root, textvariable=tStringVar2, width=40)
e2.grid(row=1, column=1)
e2.bind('<Triple-Button-1>', entry2_clear)

windnd.hook_dropfiles(e1, func=get_file_path)  # 拖拽文件

recoil_of_material_path = Path(Path.cwd()) / 'Lib' / '物料库.xlsx'
rom = openpyxl.load_workbook(recoil_of_material_path)
rom_sheet = rom[rom.sheetnames[0]]


def create():
    file_path = e1.get().replace("/", "\\")
    if_file_path_changed = False
    num = int(e2.get())
    if Path(file_path).is_file():
        wb = openpyxl.Workbook()
        ws = wb.active
        if file_path.endswith("xls"):
            if_file_path_changed = True
            file_path = xls_save_as(file_path)
            delete_file_path = Path(file_path)
        file = openpyxl.load_workbook(file_path)
        rom_row = 2
        nrom_row = 2
        sheet = file[file.sheetnames[0]]  # 打开工作表
        ws['A1'] = '物料号'
        ws['B1'] = '数量'
        ws['C1'] = '库位'
        ws['D1'] = '物料号'
        ws['E1'] = '数量'
        ws['F1'] = '库位'
        for hang in range(7, sheet.max_row - 5):
            if sheet.cell(hang, 2).value == None or sheet.cell(hang, 3).value == None:
                continue
            # 物料号
            print('第{}行的物料号是{}'.format(hang, sheet.cell(hang, 4).value), end='   ')
            true_material_no = get_material_no(sheet.cell(hang, 4).value)
            if true_material_no != None:  # 如果有物料号
                if get_material_no(sheet.cell(hang, 6).value) != None:  # 如果组成元素有物料号
                    true_material_no = get_material_no(sheet.cell(hang, 6).value)
                # 查看是否是反冲料
                print('实际物料号是{}'.format(true_material_no), end='   ')
                recoil_of_material_row = get_recoil_of_material_row(true_material_no, rom_sheet)
                print('反冲料所在行：{}'.format(recoil_of_material_row))
                if recoil_of_material_row != None:
                    ws.cell(rom_row, 1).value = true_material_no
                    # 数量
                    if get_device_count(sheet.cell(hang, 6).value) == int(sheet.cell(hang, 5).value):
                        ws.cell(rom_row, 2).value = int(sheet.cell(hang, 5).value) * num
                    else:
                        tkinter.messagebox.showinfo(title='ERROR', message='第{}行单机用量不对'.format(hang))
                        ws.cell(rom_row, 2).fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
                    ws.cell(rom_row, 3).value = 303
                    rom_row += 1
                else:
                    ws.cell(nrom_row, 4).value = true_material_no
                    # 数量
                    if get_device_count(sheet.cell(hang, 6).value) == int(sheet.cell(hang, 5).value):
                        ws.cell(nrom_row, 5).value = int(sheet.cell(hang, 5).value) * num
                    else:
                        tkinter.messagebox.showinfo(title='ERROR', message='第{}行单机用量不对'.format(hang))
                        ws.cell(nrom_row, 5).fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
                    ws.cell(nrom_row, 6).value = 104
                    nrom_row += 1
            else:
                print()
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 5
            ws.column_dimensions['C'].width = 5
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 5
            ws.column_dimensions['F'].width = 5
        wb.save(file_path.replace(Path(file_path).stem, r'DEMO'))
        if if_file_path_changed:
            Path.unlink(delete_file_path)
        tkinter.messagebox.showinfo(title='result', message='生成成功')
b1 = Button(root, text='生成', command=create)
b1.grid(row=2, column=0)

root.mainloop()
